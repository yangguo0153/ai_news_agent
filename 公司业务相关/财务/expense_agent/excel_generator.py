"""
Excel 报销单生成器
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
from dataclasses import dataclass
from datetime import datetime


@dataclass
class ExpenseItem:
    """报销条目"""
    序号: int
    日期: str
    加班时间: str       # 如 "18:00-22:00"
    出发地: str
    到达地: str
    金额: float


class ExcelGenerator:
    """Excel 报销单生成器"""
    
    def __init__(self):
        self.items: List[ExpenseItem] = []
    
    def add_item(self, item: ExpenseItem):
        """添加报销条目"""
        self.items.append(item)
    
    def add_items(self, items: List[ExpenseItem]):
        """批量添加报销条目"""
        self.items.extend(items)
    
    def generate(self, output_path: str) -> str:
        """生成 Excel 报销单"""
        if not self.items:
            raise ValueError("没有报销条目")
        
        # 转换为 DataFrame
        data = []
        for item in self.items:
            data.append({
                "序号": item.序号,
                "日期": item.日期,
                "加班时间": item.加班时间,
                "出发地": item.出发地,
                "到达地": item.到达地,
                "金额": item.金额,
            })
        
        df = pd.DataFrame(data)
        
        # 添加汇总行
        total_amount = sum(item.金额 for item in self.items)
        summary_row = pd.DataFrame([{
            "序号": "",
            "日期": "",
            "加班时间": "",
            "出发地": "",
            "到达地": "合计",
            "金额": total_amount,
        }])
        df = pd.concat([df, summary_row], ignore_index=True)
        
        # 确保输出目录存在
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 使用 openpyxl 写入并设置格式
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = Workbook()
            ws = wb.active
            ws.title = "加班打车报销单"
            
            # 添加标题行
            title = f"加班打车报销单 ({datetime.now().strftime('%Y年%m月')})"
            ws.merge_cells('A1:F1')
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # 添加表头
            headers = ["序号", "日期", "加班时间", "出发地", "到达地", "金额"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # 添加数据
            for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 3):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center' if col_idx <= 5 else 'left')
            
            # 设置列宽
            column_widths = [8, 15, 18, 20, 20, 15]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width
            
            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
                for cell in row:
                    cell.border = thin_border
            
            wb.save(str(output_path))
            
        except ImportError:
            # 如果没有 openpyxl，使用简单的 pandas 输出
            df.to_excel(str(output_path), index=False, sheet_name="加班打车报销单")
        
        return str(output_path)
    
    @property
    def total_amount(self) -> float:
        """获取总金额"""
        return sum(item.金额 for item in self.items)


if __name__ == "__main__":
    # 测试
    gen = ExcelGenerator()
    gen.add_item(ExpenseItem(
        序号=1,
        日期="2026-01-20",
        加班时间="18:00-22:00",
        出发地="公司",
        到达地="家",
        金额=35.0,
    ))
    output = gen.generate("test_report.xlsx")
    print(f"生成报销单: {output}")
