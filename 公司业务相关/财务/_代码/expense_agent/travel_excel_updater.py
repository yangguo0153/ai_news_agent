"""
差旅Excel模版更新器

更新用户提供的差旅报销Excel模版，保留原有格式
"""

import re
from pathlib import Path
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
from datetime import datetime


@dataclass
class TravelExpenseItem:
    """差旅费用条目"""
    日期: str
    天数: Optional[int] = None  # 仅第一行填写
    人员: str = ""
    性别: str = ""
    职级: str = ""
    费用名称: str = ""
    单价: float = 0.0
    数量: int = 1
    费用金额: float = 0.0
    备注: str = ""


class TravelExcelUpdater:
    """差旅Excel模版更新器"""
    
    def __init__(self, template_path: str):
        """
        :param template_path: 模版文件路径
        """
        self.template_path = Path(template_path)
        self.items: List[TravelExpenseItem] = []
        self.trip_days: int = 0  # 出差天数
        
    def set_trip_days(self, days: int):
        """设置出差天数"""
        self.trip_days = days
    
    def add_item(self, item: TravelExpenseItem):
        """添加费用条目"""
        # 自动计算金额
        if item.费用金额 == 0 and item.单价 > 0:
            item.费用金额 = item.单价 * item.数量
        self.items.append(item)
    
    def add_items(self, items: List[TravelExpenseItem]):
        """批量添加费用条目"""
        for item in items:
            self.add_item(item)
    
    def update_template(self, output_path: Optional[str] = None) -> str:
        """
        更新模版并保存
        
        :param output_path: 输出路径，默认覆盖原文件
        :return: 输出文件路径
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
        
        if output_path is None:
            output_path = str(self.template_path)
        
        # 加载模版
        wb = load_workbook(str(self.template_path))
        ws = wb.active
        
        # 找到数据起始行（跳过标题和表头）
        # 假设第1行是标题，第2行是表头，数据从第3行开始
        data_start_row = 3
        
        # 取消数据区域的所有合并单元格，避免写入错误
        self.unmerge_data_cells(ws, data_start_row)
        
        # 清空旧数据
        # 找到最后一行有数据的行 (或只清除预定义的足够多的行)
        max_row = max(ws.max_row, 30) # 至少清除到30行
        
        for row in range(data_start_row, max_row + 1):
            for col in range(1, 11):  # A-J列
                cell = ws.cell(row=row, column=col)
                # 检查是否是合并单元格 (虽然刚才unmerge了，但以防万一)
                try:
                    cell.value = None
                except AttributeError:
                    pass # 忽略合并单元格写入错误
        
        # 按费用类型排序：酒店 -> 打车 -> 餐费/礼品 -> 补助替票
        def expense_priority(item: TravelExpenseItem) -> int:
            name = item.费用名称
            if "酒店" in name or "住宿" in name:
                return 1
            elif "打车" in name or "交通" in name:
                return 2
            elif "餐" in name or "礼品" in name or "客户" in name:
                return 3
            elif "补助" in name or "替票" in name:
                return 4
            else:
                return 5
        
        sorted_items = sorted(self.items, key=expense_priority)
        
        # 写入数据
        for idx, item in enumerate(sorted_items):
            row = data_start_row + idx
            
            # 日期
            ws.cell(row=row, column=1, value=item.日期)
            
            # 天数（仅第一行填写）
            if idx == 0 and self.trip_days > 0:
                ws.cell(row=row, column=2, value=self.trip_days)
            
            # 人员
            ws.cell(row=row, column=3, value=item.人员)
            
            # 性别
            ws.cell(row=row, column=4, value=item.性别)
            
            # 职级
            ws.cell(row=row, column=5, value=item.职级)
            
            # 费用名称
            ws.cell(row=row, column=6, value=item.费用名称)
            
            # 单价
            ws.cell(row=row, column=7, value=item.单价)
            
            # 数量
            ws.cell(row=row, column=8, value=item.数量)
            
            # 费用金额
            ws.cell(row=row, column=9, value=item.费用金额)
            
            # 备注
            ws.cell(row=row, column=10, value=item.备注)
            
            # 设置居中对齐
            for col in range(1, 11):
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal='center', 
                    vertical='center'
                )
        
        # 添加合计行
        if sorted_items:
            total_row = data_start_row + len(sorted_items)
            ws.cell(row=total_row, column=6, value="合计")
            total_amount = sum(item.费用金额 for item in sorted_items)
            ws.cell(row=total_row, column=9, value=total_amount)
            
            # 设置合计行格式
            for col in range(1, 11):
                ws.cell(row=total_row, column=col).alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )
        
        # 保存
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        
        return str(output_path)
    
    def unmerge_data_cells(self, ws, start_row: int):
        """取消数据区域的合并单元格"""
        from openpyxl.utils import range_boundaries
        
        # 获取所有合并单元格范围
        merged_ranges = list(ws.merged_cells.ranges)
        
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            
            # 如果合并范围涉及数据行 (start_row及以下)，则取消合并
            if max_row >= start_row:
                try:
                    ws.unmerge_cells(str(merged_range))
                except:
                    pass
    
    @property
    def total_amount(self) -> float:
        """获取总金额"""
        return sum(item.费用金额 for item in self.items)


def parse_date_to_display(date_str: str) -> str:
    """
    将日期字符串转换为显示格式
    
    输入: 2026-01-25 或 2026年1月25日
    输出: 1月25日
    """
    # 尝试解析不同格式
    patterns = [
        (r'(\d{4})-(\d{1,2})-(\d{1,2})', lambda m: f"{int(m.group(2))}月{int(m.group(3))}日"),
        (r'(\d{4})年(\d{1,2})月(\d{1,2})日', lambda m: f"{int(m.group(2))}月{int(m.group(3))}日"),
    ]
    
    for pattern, formatter in patterns:
        match = re.match(pattern, date_str)
        if match:
            return formatter(match)
    
    return date_str


if __name__ == "__main__":
    # 测试
    print("差旅Excel更新器模块加载成功")
    
    updater = TravelExcelUpdater("/tmp/test_template.xlsx")
    print(f"模版路径: {updater.template_path}")
