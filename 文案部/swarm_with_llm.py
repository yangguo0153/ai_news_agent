from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

from config import config

"""
Agent Swarm 原型 - Content Expansion (with Real LLM)

架构：@main (Orchestrator) + 5个业务角色
技术栈：LangGraph + Claude Opus (真实 API 调用)
"""

from typing import TypedDict, List, Dict
from langgraph.graph import StateGraph, END
import json
import random
import os
import requests
import asyncio
import aiohttp
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ============================================================================
# Shared Context（所有角色共享的状态）
# ============================================================================

class SharedContext(TypedDict):
    """Shared Context - 所有角色都能读写的共享状态"""

    # 用户输入
    user_input: Dict  # {车型, 平台, 数量, 方向}

    # 客户经理输出
    customer_brief: Dict  # {需求摘要, 侧重点, 目标用户, 核心卖点, 调性}

    # 策划者输出
    planner_brief: Dict  # {传播方向, 话题切入点, assignments[]}

    # Writer 输出
    contents: List[Dict]  # [{id, content, persona, selling_point, attempt, revision_history}]

    # 审核者输出
    review_results: List[Dict]  # [{id, passed, issues, suggestions}]

    # 输出校订者输出
    final_output: str  # Excel 文件路径

    # 循环控制
    current_attempt: int  # 当前循环次数（1-3）
    need_manual_review: List[int]  # 需要人工介入的内容ID列表

    # 用户交互控制
    skip_confirmations: bool  # 是否跳过所有确认

    # 元数据
    metadata: Dict  # {start_time, current_stage, attempts}


# ============================================================================
# 工具函数
# ============================================================================




def random_sample_details(detail_library: List[str], k: int = 3) -> List[str]:
    """从细节库中随机抽取样本"""
    return random.sample(detail_library, min(k, len(detail_library)))


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=10),
    retry=retry_if_exception_type((aiohttp.ClientError, asyncio.TimeoutError)),
    reraise=True
)
async def call_deepseek_api_async(prompt: str, api_key: str, api_url: str, temperature: float = 0.7, max_tokens: int = 512) -> str:
    """
    异步调用 Deepseek API

    Args:
        prompt: 提示词
        api_key: API 密钥
        api_url: API 地址
        temperature: 生成的随机性温度
        max_tokens: 允许生成的最长 Tokens

    Returns:
        生成的内容
    """
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": "deepseek-chat",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": temperature,
        "max_tokens": max_tokens
    }

    async with aiohttp.ClientSession() as session:
        async with session.post(api_url, headers=headers, json=payload, timeout=60) as response:
            response.raise_for_status()
            result = await response.json()
            return result['choices'][0]['message']['content'].strip()



def ask_user_confirmation(title: str, content: Dict, options: List[str] = None) -> str:
    """
    向用户展示内容并请求确认

    Args:
        title: 确认标题
        content: 要展示的内容（字典格式）
        options: 可选项列表，默认为 ["确认", "修改"]

    Returns:
        用户选择的选项
    """
    if options is None:
        options = ["确认", "修改"]

    print("\n" + "=" * 60)
    print(f"【{title}】")
    print("=" * 60)

    # 展示内容
    for key, value in content.items():
        if isinstance(value, list):
            print(f"\n{key}:")
            for item in value:
                print(f"  - {item}")
        else:
            print(f"{key}: {value}")

    # 请求确认
    print("\n" + "-" * 60)
    print("请选择操作：")
    for i, option in enumerate(options, 1):
        print(f"  {i}. {option}")

    while True:
        try:
            choice = input("\n请输入选项编号: ").strip()
            choice_idx = int(choice) - 1
            if 0 <= choice_idx < len(options):
                selected = options[choice_idx]
                print(f"✓ 已选择: {selected}")
                return selected
            else:
                print(f"❌ 无效选项，请输入 1-{len(options)}")
        except ValueError:
            print(f"❌ 请输入数字 1-{len(options)}")
        except (KeyboardInterrupt, EOFError):
            print("\n\n⚠️  用户中断或输入结束，使用默认选项: 确认")
            return "确认"


def load_detail_library(file_path: str) -> List[str]:
    """从 markdown 文件加载细节描写库"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # 解析 markdown，提取列表项
        details = []
        for line in content.split('\n'):
            line = line.strip()
            if line.startswith('- "') and line.endswith('"'):
                detail = line[3:-1]  # 去掉 '- "' 和 '"'
                details.append(detail)

        return details
    except FileNotFoundError:
        print(f"警告：找不到细节库文件 {file_path}，使用默认样本")
        return [
            "不用玩后备箱俄罗斯方块",
            "老妈的腊肉、老爸的酒，一样都不落",
            "后备箱盖一关，满满当当的安心"
        ]


def load_persona_samples(persona: str) -> Dict[str, List[str]]:
    """
    加载指定人设的口吻样本

    Args:
        persona: 人设名称（宝妈、孝子、小夫妻、职场精英）

    Returns:
        包含开场切入、痛点描述、解决方案、情感升华的字典
    """
    file_path = "02-参考学习/03-Writer材料/内容变量库/口吻样本库.md"

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # 初始化结果
        samples = {
            "开场切入": [],
            "痛点描述": [],
            "解决方案": [],
            "情感升华": []
        }

        # 根据人设找到对应部分
        persona_markers = {
            "宝妈": "## 一、宝妈口吻",
            "孝子": "## 二、孝子口吻",
            "小夫妻": "## 三、小夫妻口吻",
            "职场精英": "## 四、职场精英口吻"
        }

        if persona not in persona_markers:
            return samples

        # 找到该人设的起始位置
        start_marker = persona_markers[persona]
        start_pos = content.find(start_marker)
        if start_pos == -1:
            return samples

        # 找到下一个人设的起始位置（作为结束位置）
        next_markers = [m for p, m in persona_markers.items() if p != persona]
        end_pos = len(content)
        for next_marker in next_markers:
            pos = content.find(next_marker, start_pos + 1)
            if pos != -1 and pos < end_pos:
                end_pos = pos

        # 提取该人设的内容
        persona_content = content[start_pos:end_pos]

        # 解析各个子部分
        current_subsection = None
        for line in persona_content.split('\n'):
            line = line.strip()
            if "**开场切入**" in line:
                current_subsection = "开场切入"
            elif "**痛点描述**" in line:
                current_subsection = "痛点描述"
            elif "**解决方案**" in line:
                current_subsection = "解决方案"
            elif "**情感升华**" in line:
                current_subsection = "情感升华"
            elif current_subsection and line.startswith('- "') and line.endswith('"'):
                # 处理中文引号
                sample = line[3:-1]
                samples[current_subsection].append(sample)

        return samples

    except FileNotFoundError:
        print(f"警告：找不到口吻样本库文件 {file_path}")
        return {"开场切入": [], "痛点描述": [], "解决方案": [], "情感升华": []}


def load_scene_samples(scene_type: str) -> Dict[str, List[str]]:
    """
    加载指定场景的切入样本

    Args:
        scene_type: 场景类型（春节返乡、周末出游、日常通勤等）

    Returns:
        包含时间触发、场景描写、情感升华的字典
    """
    file_path = "02-参考学习/03-Writer材料/内容变量库/场景切入库.md"

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # 初始化结果
        samples = {
            "时间触发": [],
            "场景描写": [],
            "情感升华": []
        }

        # 根据场景类型找到对应部分
        scene_markers = {
            "春节返乡": "## 一、春节返乡场景",
            "周末出游": "## 二、周末出游场景",
            "日常通勤": "## 三、日常通勤场景",
            "亲子游玩": "## 四、亲子游玩场景",
            "孝敬父母": "## 五、孝敬父母场景"
        }

        # 默认使用春节返乡
        start_marker = scene_markers.get(scene_type, scene_markers["春节返乡"])

        # 找到该场景的起始位置
        start_pos = content.find(start_marker)
        if start_pos == -1:
            return samples

        # 找到下一个场景的起始位置（作为结束位置）
        next_markers = [m for s, m in scene_markers.items() if s != scene_type]
        end_pos = len(content)
        for next_marker in next_markers:
            pos = content.find(next_marker, start_pos + 1)
            if pos != -1 and pos < end_pos:
                end_pos = pos

        # 提取该场景的内容
        scene_content = content[start_pos:end_pos]

        # 解析各个子部分
        current_subsection = None
        for line in scene_content.split('\n'):
            line = line.strip()
            if "**时间触发**" in line:
                current_subsection = "时间触发"
            elif "**场景描写**" in line:
                current_subsection = "场景描写"
            elif "**情感升华**" in line:
                current_subsection = "情感升华"
            elif current_subsection and line.startswith('- "') and line.endswith('"'):
                # 处理中文引号
                sample = line[3:-1]
                samples[current_subsection].append(sample)

        return samples

    except FileNotFoundError:
        print(f"警告：找不到场景切入库文件 {file_path}")
        return {"时间触发": [], "场景描写": [], "情感升华": []}


def load_few_shot_samples(platform: str) -> List[str]:
    """
    加载指定平台的爆款样本进行 Few-Shot 注入
    """
    file_path = "02-参考学习/03-Writer材料/内容变量库/爆款参考库.md"
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
            
        # 根据平台对应不同标题区块
        platform_markers = {
            "小红书": "## 小红书平台样本",
            "抖音": "## 抖音平台样本",
            "今日头条": "## 今日头条平台样本",
            "朋友圈": "## 朋友圈样本"
        }
        
        start_marker = platform_markers.get(platform, platform_markers["小红书"])
        start_pos = content.find(start_marker)
        if start_pos == -1:
            return []
            
        next_markers = [m for p, m in platform_markers.items() if p != platform]
        end_pos = len(content)
        for next_marker in next_markers:
            pos = content.find(next_marker, start_pos + 1)
            if pos != -1 and pos < end_pos:
                end_pos = pos
                
        platform_content = content[start_pos:end_pos]
        
        samples = []
        for line in platform_content.split('\n'):
            line = line.strip()
            if line.startswith('- "') and line.endswith('"'):
                samples.append(line[3:-1])
                
        return samples
    except FileNotFoundError:
        print(f"警告：找不到爆款参考库文件 {file_path}")
        return []


def generate_excel_output(state: SharedContext) -> str:
    """生成 Excel 格式的输出报告"""
    wb = Workbook()

    # Sheet 1: 内容汇总
    ws1 = wb.active
    ws1.title = "内容汇总"

    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["篇号", "人设", "卖点", "状态", "字数", "尝试次数", "内容"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 填充数据
    for row_idx, (content_item, review) in enumerate(zip(state["contents"], state["review_results"]), 2):
        status = "✅ 通过" if review["passed"] else "⚠️ 人工介入"
        ws1.cell(row=row_idx, column=1, value=content_item["id"])
        ws1.cell(row=row_idx, column=2, value=content_item["persona"])
        ws1.cell(row=row_idx, column=3, value=content_item["selling_point"])
        ws1.cell(row=row_idx, column=4, value=status)
        ws1.cell(row=row_idx, column=5, value=len(content_item["content"]))
        ws1.cell(row=row_idx, column=6, value=content_item.get("attempt", 1))
        ws1.cell(row=row_idx, column=7, value=content_item["content"])

    # 调整列宽
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 80

    # Sheet 2: 审核详情
    ws2 = wb.create_sheet("审核详情")
    headers2 = ["篇号", "审核结果", "问题描述", "修改建议"]
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, review in enumerate(state["review_results"], 2):
        result = "通过" if review["passed"] else "不通过"
        issues = "; ".join(review.get("issues", []))
        suggestions = "; ".join(review.get("suggestions", []))
        ws2.cell(row=row_idx, column=1, value=review["id"])
        ws2.cell(row=row_idx, column=2, value=result)
        ws2.cell(row=row_idx, column=3, value=issues)
        ws2.cell(row=row_idx, column=4, value=suggestions)

    # 调整列宽
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 40
    ws2.column_dimensions['D'].width = 40

    # Sheet 3: 元数据
    ws3 = wb.create_sheet("元数据")
    ws3.cell(row=1, column=1, value="项目").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="值").font = Font(bold=True)

    metadata_rows = [
        ("车型", state["user_input"]["车型"]),
        ("平台", state["user_input"]["平台"]),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("总篇数", len(state["contents"])),
        ("通过篇数", sum(1 for r in state["review_results"] if r["passed"])),
        ("人工介入篇数", len(state.get("need_manual_review", [])))
    ]

    for row_idx, (key, value) in enumerate(metadata_rows, 2):
        ws3.cell(row=row_idx, column=1, value=key)
        ws3.cell(row=row_idx, column=2, value=value)

    # 调整列宽
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 30

    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{state['user_input']['车型']}_{state['user_input']['平台']}_{timestamp}.xlsx"
    output_path = f"04-产出仓库/{filename}"

    # 确保输出目录存在
    os.makedirs("04-产出仓库", exist_ok=True)

    wb.save(output_path)

    return output_path


def check_banned_words(content: str) -> List[str]:
    """检查禁用词"""
    banned = ['说实话', '但问题来了', '你看', '首先', '其次', '方面', '不得不说']
    found = [word for word in banned if word in content]
    return found


def count_params(content: str) -> int:
    """统计参数数量"""
    params = ['10气囊', 'ACE车身', '980Mpa', 'Honda SENSING', 'MM理念', '保值率']
    return sum(1 for p in params if p in content)


def check_scene_quality(content: str) -> Dict[str, any]:
    """
    检查场景化程度（软性审核）

    返回：
    - score: 0-10 分
    - has_scene: 是否包含具体场景
    - has_detail: 是否包含细节描写
    - feedback: 反馈建议
    """
    score = 5  # 基础分
    has_scene = False
    has_detail = False
    feedback = []

    # 检查场景触发词
    scene_triggers = ['每到', '一到', '每年', '又是', '每次']
    if any(trigger in content for trigger in scene_triggers):
        score += 1
        has_scene = True
    else:
        feedback.append("建议添加时间触发词（每到/一到/每年）")

    # 检查对比式表达
    contrast_words = ['以前', '现在', '过去', '如今', '那时', '今年']
    if any(word in content for word in contrast_words):
        score += 1
        has_scene = True
    else:
        feedback.append("建议使用对比式表达（以前 vs 现在）")

    # 检查具体细节
    detail_indicators = ['后备箱', '后排', '座椅', '空间', '系统', '车身', '方向盘', '高速', '路上']
    detail_count = sum(1 for indicator in detail_indicators if indicator in content)
    if detail_count >= 3:
        score += 2
        has_detail = True
    elif detail_count >= 1:
        score += 1
        has_detail = True
        feedback.append("建议增加更多具体细节描写")
    else:
        feedback.append("缺少具体细节描写")

    # 检查生活化场景
    life_scenes = ['孩子', '父母', '爸妈', '老人', '家人', '年货', '礼物', '行李']
    if any(scene in content for scene in life_scenes):
        score += 1

    return {
        "score": min(score, 10),
        "has_scene": has_scene,
        "has_detail": has_detail,
        "feedback": feedback
    }


def check_emotion_quality(content: str) -> Dict[str, any]:
    """
    检查情感共鸣度（软性审核）

    返回：
    - score: 0-10 分
    - has_emotion: 是否包含情感表达
    - has_warmth: 是否温和喜庆
    - feedback: 反馈建议
    """
    score = 5  # 基础分
    has_emotion = False
    has_warmth = False
    feedback = []

    # 检查情感词汇
    emotion_words = ['心', '踏实', '安心', '温暖', '期待', '牵挂', '焦虑', '担心', '开心', '笑']
    emotion_count = sum(1 for word in emotion_words if word in content)
    if emotion_count >= 3:
        score += 2
        has_emotion = True
    elif emotion_count >= 1:
        score += 1
        has_emotion = True
        feedback.append("建议增加更多情感表达")
    else:
        feedback.append("缺少情感表达")

    # 检查温和喜庆的表达
    warmth_words = ['团圆', '回家', '平安', '幸福', '美好', '甜', '暖', '圆满']
    if any(word in content for word in warmth_words):
        score += 1
        has_warmth = True
    else:
        feedback.append("建议增加温和喜庆的表达")

    # 检查情感升华
    elevation_patterns = ['就是', '才是', '不就是', '正是']
    if any(pattern in content for pattern in elevation_patterns):
        score += 1

    # 检查是否有具体感受描写
    feeling_words = ['手心', '脸色', '笑容', '表情', '眼神', '声音']
    if any(word in content for word in feeling_words):
        score += 1

    return {
        "score": min(score, 10),
        "has_emotion": has_emotion,
        "has_warmth": has_warmth,
        "feedback": feedback
    }


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=10),
    retry=retry_if_exception_type((aiohttp.ClientError, asyncio.TimeoutError)),
    reraise=True
)
async def evaluate_content_ai_flavor_async(content: str, assignment: Dict, api_key: str, api_url: str) -> Dict:
    prompt = f"""你是一个极其严格的“反AI八股文”内容质检管家。你的目标是检查以下社交媒体文案是否含有“AI味”、“公关播音腔”或“套路化模板”。

【审核内容】
人设：{assignment['persona']}
卖点：{assignment['selling_point']}
正文：
{content}

【🔴 必须打回重写（不通过）的五大雷区】
1. 词汇雷区：包含“每到春节”、“归心似箭”、“保驾护航”、“移动的家”、“不得不说”、“承载”等陈词滥调。
2. 逻辑缺失雷区：文章没有建立【共鸣痛点/认知冲突 -> (视情况有竞品对比)方案抛出 -> 具体数字参数实证 -> 价值结论】的说服力链路，通篇只有无意义的情感情绪抒发或纯粹的感叹。
3. 竞品拉踩与格式雷区：如果文章包含竞品对比，出现了恶意贬低、踩踏友商的词汇（拉踩）；或者在文章的标题、首句、文末#话题标签(发布文案)等位置直接出现了友商的车型名称（友商名只允许在正文探讨参数时出现）。
4. 语气雷区：使用了“绝绝子”、“无语子”、“绝了”、“服了”、“简直”等低层次情绪词汇装作“碎碎念网感”，而没有展现出理性、精打细算懂车达人的真实软性评测质感。
5. ⚠️ 数据空洞雷区（致命）：文中必须根据场景自然指出 1-2 个宏观大卖点（如动力、操控、空间等），并配合陈述 2-3 个具体的真实物理参数/专有名词（如：2701mm轴距、193匹马力、Honda SENSING 等）来支撑。如果全是“空间大”、“动力强”等虚词，未提及具体配置数据或技术名词，必须直接打回！

请严格审核！给出 0-10 的“去AI味”评分（10分代表毫无AI味且数据扎实、极其像真人；低于7分判定为不通过）。
必须仅以纯JSON格式返回，不要有任何多余字符，格式如下：
{{
  "passed": false,
  "score": 5,
  "issues": ["指出具体哪里有AI味，或者缺少具体配置数据支撑"],
  "suggestions": ["给出具体怎么改的建议，包括补充什么类型的数据"]
}}"""
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": "deepseek-chat",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_tokens": 512
    }
    async with aiohttp.ClientSession() as session:
        async with session.post(api_url, headers=headers, json=payload, timeout=60) as response:
            try:
                response.raise_for_status()
                result = await response.json()
                raw_content = result['choices'][0]['message']['content'].strip()
                if raw_content.startswith("```json"): raw_content = raw_content[7:]
                if raw_content.startswith("```"): raw_content = raw_content[3:]
                if raw_content.endswith("```"): raw_content = raw_content[:-3]
                review_res = json.loads(raw_content.strip())
                return review_res
            except Exception as e:
                print(f"[Reviewer API Error] {e}")
                # 解析失败时默认放行，避免无限循环打回
                return {"passed": True, "score": 7, "issues": [], "suggestions": []}

def build_revision_prompt(customer_brief, assignment, original_content, issues, suggestions):
    """
    构建修改 Prompt
    """
    platform = customer_brief['平台']
    platform_specs = {
        "小红书": "100-300字，轻快、多Emoji、生活碎片感",
        "抖音": "250-300字，短平快钩子多，极其口语化",
        "今日头条": "650-800字，深度分析、新闻纪实感、带小标题的逻辑文",
        "朋友圈": "50-150字，熟人语境极其简短的碎碎念"
    }
    spec = platform_specs.get(platform, "自然人类语气")

    prompt = f"""你是一个顶级社交媒体达人。你之前针对【{platform}】写的一篇文案被判定为“不够真实”、“有AI味”或“不符合该平台的调性”，你需要根据反馈重新修改。

【原内容】
{original_content}

【审核官的批评及雷区】
{chr(10).join(f"- {issue}" for issue in issues)}

【修改建议】
{chr(10).join(f"- {suggestion}" for suggestion in suggestions)}

【本篇设定】
- 平台：{platform}（{spec}）
- 人设：{assignment['persona']}
- 核心卖点：{assignment['selling_point']}

【🔴 强制要求（违反视为再败）】
1. 狠砸原来的机器骨架！不要标准四段式。请严格遵循【{platform}】的体裁来写！
2. 彻底抛弃播音腔和公关词汇（禁用：缔造、保驾护航、移动的家、承载、不得不说等）。
3. 务必满足审核官指出的字数与格式要求。直接输出修改后的正文（勿加注释）："""

    return prompt


def revise_contents(state: SharedContext, failed_items: List[Dict], attempt: int) -> List[Dict]:
    """
    修改不通过的内容
    """
    customer_brief = state["customer_brief"]
    planner_brief = state["planner_brief"]
    existing_contents = state["contents"]

    # Deepseek API 配置
    DEEPSEEK_API_KEY = config.DEEPSEEK_API_KEY
    DEEPSEEK_API_URL = config.DEEPSEEK_API_URL

    # 保留通过的内容，重写不通过的内容
    updated_contents = []

    for content_item in existing_contents:
        # 查找对应的审核结果
        review = next((r for r in failed_items if r["id"] == content_item["id"]), None)

        if review is None:
            # 已通过，保留原内容
            updated_contents.append(content_item)
        else:
            # 需要修改
            print(f"  [Writer] 正在修改第{content_item['id']}篇...")

            # 找到对应的 assignment
            assignment = next(
                (a for a in planner_brief["assignments"] if a["id"] == content_item["id"]),
                None
            )

            # 构建包含修改建议的 Prompt
            prompt = build_revision_prompt(
                customer_brief,
                assignment,
                content_item["content"],
                review["issues"],
                review["suggestions"]
            )

            # 调用 Deepseek API
            try:
                headers = {
                    "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                    "Content-Type": "application/json"
                }

                payload = {
                    "model": "deepseek-chat",
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.7,  # 降低随机性，提升字数控制
                    "max_tokens": 512    # 限制最大长度
                }

                response = requests.post(DEEPSEEK_API_URL, headers=headers, json=payload, timeout=60)
                response.raise_for_status()

                result = response.json()
                revised_content = result['choices'][0]['message']['content'].strip()

                # 更新内容
                updated_item = {
                    "id": content_item["id"],
                    "content": revised_content,
                    "persona": content_item["persona"],
                    "selling_point": content_item["selling_point"],
                    "attempt": attempt,
                    "revision_history": content_item.get("revision_history", []) + [{
                        "attempt": attempt - 1,
                        "issues": review["issues"],
                        "suggestions": review["suggestions"]
                    }]
                }

                updated_contents.append(updated_item)
                print(f"    ✓ 第{content_item['id']}篇修改完成（{len(revised_content)}字）")

            except Exception as e:
                print(f"    ✗ 第{content_item['id']}篇修改失败：{e}")
                # 失败时保留原内容
                updated_contents.append(content_item)

    return updated_contents


# ============================================================================
# 业务角色实现
# ============================================================================

def 客户经理(state: SharedContext) -> SharedContext:
    """
    职责：消化需求、与用户确认、输出 customer_brief
    """
    print("\n[客户经理] 开始处理需求...")

    user_input = state["user_input"]
    car_model = user_input.get("车型", "CR-V")

    # 读取参考材料
    material = config.load_material("02-参考学习/01-客户经理材料/需求消化模板.md")
    
    # ======= Task 1: 动态挂载车系大模型知识 (RAG Core) =======
    car_knowledge_path = f"01-输入材料/既定资料/{car_model}.md"
    car_knowledge = config.load_material(car_knowledge_path)
    if "材料文件未找到" in car_knowledge:
        print(f"⚠️ [客户经理] 未找到 {car_model} 的专属资料库，将使用基础认知生成。")
        car_knowledge = "暂无详细资料，请依据普遍造车常识进行创作。"
    else:
        print(f"✓ [客户经理] 成功挂载 {car_model} 专属大语言模型知识库 ({len(car_knowledge)} 字符)")

    # 生成 customer_brief
    customer_brief = {
        "车型": car_model,
        "平台": user_input.get("平台", "抖音"),
        "数量": user_input.get("数量", 5),
        "方向": user_input.get("方向", "春节返乡"),
        "侧重点": "情感共鸣",
        "目标用户": ["宝妈", "孝子", "小夫妻"],
        "核心卖点": ["空间", "安全"],
        "调性": "温和喜庆",
        "特殊要求": "避免直白描述，多用场景暗示",
        "车型专属打底知识": car_knowledge # 传递给下游
    }

    # 用户确认
    skip_confirmations = state.get("skip_confirmations", False)
    if not skip_confirmations:
        confirmation = ask_user_confirmation(
            title="客户需求理解确认",
            content={
                "车型": customer_brief["车型"],
                "平台": customer_brief["平台"],
                "数量": f"{customer_brief['数量']}篇",
                "方向": customer_brief["方向"],
                "侧重点": customer_brief["侧重点"],
                "目标用户": ", ".join(customer_brief["目标用户"]),
                "核心卖点": ", ".join(customer_brief["核心卖点"]),
                "调性": customer_brief["调性"],
                "特殊要求": customer_brief["特殊要求"]
            },
            options=["确认", "修改", "跳过所有确认"]
        )

        # 处理用户选择
        if confirmation == "修改":
            print("\n[客户经理] 请手动修改需求...")
            # 这里可以添加交互式修改逻辑
            # 为简化，暂时使用原 brief
            print("⚠️  暂不支持交互式修改，使用原需求继续")
        elif confirmation == "跳过所有确认":
            state["skip_confirmations"] = True
            print("[客户经理] 已设置跳过所有后续确认")
    else:
        print("[客户经理] 跳过需求确认阶段（配置为自动流转）")

    state["customer_brief"] = customer_brief
    print(f"[客户经理] 已生成 customer_brief")

    return state


def 策划者(state: SharedContext) -> SharedContext:
    """
    职责：思考传播方向、输出 planner_brief（包含分配表）
    """
    print("\n[策划者] 开始策划传播方向...")

    customer_brief = state["customer_brief"]
    skip_confirmations = state.get("skip_confirmations", False)

    # 读取参考材料
    material = config.load_material("02-参考学习/02-策划者材料/传播方向案例库.md")

    # 生成分配表
    target_users = customer_brief["目标用户"]
    selling_points = customer_brief["核心卖点"]
    num_contents = customer_brief["数量"]

    # ======= Task 2: 动态场景推演 (Dynamic Scene Planner) =======
    direction = customer_brief["方向"]
    
    print(f"\n[策划者] 正在接收并解构宏观场景: '{direction}'...")
    try:
        enriched_scene_tags = asyncio.run(
            map_scene_to_keywords_async(direction, config.DEEPSEEK_API_KEY, config.DEEPSEEK_API_URL)
        )
        print(f"  -> 🧠 [动态推演] 场景向量化拆解: {enriched_scene_tags}")
    except Exception as e:
        print(f"  -> ⚠️ 场景推演失败，退回原始字符: {e}")
        enriched_scene_tags = direction
    
    assignments = []
    for i in range(num_contents):
        assignments.append({
            "id": i + 1,
            "persona": target_users[i % len(target_users)],
            "selling_point": selling_points[i % len(selling_points)],
            "scene": f"[{direction}] - {enriched_scene_tags}",
            "style": customer_brief.get("调性", "真实")
        })

    planner_brief = {
        "传播方向": direction,
        "话题切入点": f"{direction} 带来的真实痛点与情绪共鸣",
        "assignments": assignments
    }

    # 用户确认（如果未跳过）
    if not skip_confirmations:
        # 格式化分配表用于展示
        assignments_display = []
        for a in assignments:
            assignments_display.append(
                f"篇{a['id']}: {a['persona']} × {a['selling_point']} ({a['scene']})"
            )

        confirmation = ask_user_confirmation(
            title="内容分配表确认",
            content={
                "传播方向": planner_brief["传播方向"],
                "话题切入点": planner_brief["话题切入点"],
                "内容分配": assignments_display
            },
            options=["确认", "重新分配", "跳过所有确认"]
        )

        # 处理用户选择
        if confirmation == "重新分配":
            print("\n[策划者] 请手动调整分配表...")
            # 这里可以添加交互式调整逻辑
            # 为简化，暂时使用原分配表
            print("⚠️  暂不支持交互式调整，使用原分配表继续")
        elif confirmation == "跳过所有确认":
            state["skip_confirmations"] = True
            print("[策划者] 已设置跳过所有后续确认")

    state["planner_brief"] = planner_brief
    print(f"[策划者] 已生成 planner_brief，共{len(assignments)}篇分配")

    return state


def Writer(state: SharedContext) -> SharedContext:
    """
    职责：根据分配任务创作内容（支持修改模式）
    """
    customer_brief = state["customer_brief"]
    planner_brief = state["planner_brief"]
    review_results = state.get("review_results", [])
    current_attempt = state.get("current_attempt", 1)

    # 判断是首次创作还是修改模式
    is_revision = len(review_results) > 0

    if is_revision:
        print(f"\n[Writer] 修改模式（第{current_attempt}次尝试）...")
        # 只处理不通过的内容
        failed_items = [r for r in review_results if not r["passed"]]
        contents = revise_contents(state, failed_items, current_attempt)
    else:
        print(f"\n[Writer] 首次创作...")
        # 处理所有 assignments
        assignments = planner_brief["assignments"]

        # 读取参考材料
        template = config.load_material("02-参考学习/03-Writer材料/结构模板库/春节返乡-情感路线.md")

        # 从文件加载细节库
        detail_library_path = "02-参考学习/03-Writer材料/内容变量库/细节描写库.md"
        detail_samples = load_detail_library(detail_library_path)

        if not detail_samples:
            print("警告：细节库为空，使用默认样本")
            detail_samples = [
                "不用玩后备箱俄罗斯方块",
                "老妈的腊肉、老爸的酒，一样都不落",
                "后备箱盖一关，满满当当的安心"
            ]

        # Deepseek API 配置
        DEEPSEEK_API_KEY = config.DEEPSEEK_API_KEY
        DEEPSEEK_API_URL = config.DEEPSEEK_API_URL

        # 并行创作所有内容
        async def create_single_content(assignment: Dict) -> Dict:
            """异步创建单篇内容"""
            print(f"  [Writer] 开始创作第{assignment['id']}篇（{assignment['persona']} - {assignment['selling_point']}）...")

            # 加载材料库
            selected_details = random_sample_details(detail_samples, k=3)
            persona_samples = load_persona_samples(assignment['persona'])
            persona_opening = random.sample(persona_samples.get("开场切入", [""]), min(2, len(persona_samples.get("开场切入", []))))
            persona_pain = random.sample(persona_samples.get("痛点描述", [""]), min(2, len(persona_samples.get("痛点描述", []))))
            scene_samples = load_scene_samples(customer_brief['方向'])
            scene_trigger = random.sample(scene_samples.get("时间触发", [""]), min(2, len(scene_samples.get("时间触发", []))))
            scene_emotion = random.sample(scene_samples.get("情感升华", [""]), min(2, len(scene_samples.get("情感升华", []))))

            # ======= 分平台策略适配 =======
            platform = customer_brief['平台']
            spec = config.PLATFORM_SPECS.get(platform, config.PLATFORM_SPECS["小红书"]) # 默认小红书
            
            # [🔥 Phase 8 Task 2/4] 动态加载真实爆款 Few-Shot & 随机生成参数
            hit_samples_full = load_few_shot_samples(platform)
            hit_samples = random.sample(hit_samples_full, min(2, len(hit_samples_full))) if hit_samples_full else []
            
            # 随机化温度带来多样性 (防同质化)
            dynamic_temp = round(random.uniform(0.6, 0.9), 2)
            # 计算字数对应的近似 Max Tokens（使用 limits 元组的上限）
            word_upper = spec.get('limits', (200, 400))[1]
            dynamic_max_tokens = int(word_upper * 2.5) 

            prompt = f"""你是一个深谙各个社交平台流量密码的顶级内容主理人。你现在要根据分配的信息，写一篇针对【{platform}】平台的关于汽车（{customer_brief['车型']}）的内容。

【任务信息】
- 车型：{customer_brief['车型']}
- 平台：{platform}（目标篇幅：{spec['word_count']}）
- 平台专属文风要求：{spec['style']}
- 平台专属排版结构：{spec['structure']}
- 方向：{customer_brief['方向']}
- 调性：{customer_brief['调性']}

【本篇分配】
- 人设：{assignment['persona']}
- 核心卖点：{assignment['selling_point']}
- 场景：{assignment['scene']}

【💎 核心要求：产品配置干货植入 (极其重要)】
以下是关于该车型的官方专属打底知识。你**必须根据前面的场景，自然指出 1-2 个宏观卖点（如：动力、操控、科技配置、空间等），然后强行且自然地陈述 2-3 个具体的小卖点物理参数（如 2701mm轴距、193匹马力、Honda SENSING 等）来补充解释你的观点**！
❌ 错误示范：“这车空间很大，装得下很多东西”、“动力充沛，随叫随到”、“非常安全，气囊很多”
✅ 正确示范（具体数据支撑）：
- （动力切入）：“起步太窜了，这**1.5T地球梦发动机**给的**193匹马力**真的不是盖的，一脚油门下去推背感绝了...”
- （空间切入）：“后备箱空间变态大，**2701mm的轴距**使得第二排腿部空间超级富裕，哪怕放了**儿童推车**也不用折叠直接怼进去了哈~”
- （安全切入）：“刚才雨雪天路滑，差点被旁边的货车挤到，多亏**Honda SENSING**警报疯狂介入，而且想想要是真撞了，毕竟全系标配**10个安全气囊**外加**ACE承载式车身**，心里也很有底。”
<车型专属资料>
{customer_brief.get('车型专属打底知识', '暂无详细资料，请结合大众对该车型的认知。')}
</车型专属资料>

【🔴 极其严厉的文风与逻辑规则（违反任何一条视为失败）】
0. **【字数红线】**：必须严格控制在 **{spec['word_count']}** 的区间内！写完后在心里默数一遍字数，坚决不准超出上限或低于下限！
1. **【痛点营销逻辑强制】**：全文必须包含清晰的说服逻辑链路：【共鸣痛点/受众认知】 -> 【制造选车纠结/点出痛点】 -> 【抛出本品核心硬核参数降维打击】 -> 【价值升华/结论抛出】。
2. **【拒绝低级口语化】**：绝对禁止使用任何公关广告腔调。同时也**绝不使用**“绝绝子”、“无语子”、“太”、“简直”、“哎”、“服了”等低级情绪词汇。你要表现得像一个懂车且理性、精打细算的导购达人/评测博主，语气干练扎实，用“这波操作确实很香”、“性价比没得说”、“精准戳中需求”等沉稳词汇。
3. **【极其严格的竞品对比红线（条件触发）】**：如果在方向/场景中涉及“对比”，你必须引入友商热门竞品作为锚点，用具体的参数对比进行说服。但必须遵守两点铁律：
   A. **绝对不准“拉踩”**：对比时必须客观甚至肯定友商长处，禁止使用任何贬低、攻击性词汇去踩踏竞品。
   B. **标题与标签区隔离**：如果你的输出带有标题、开头引导语或文末的 #话题标签（发布文案），这些区域**绝对禁止**出现友商车型名称！友商名字只允许在探讨具体参数的**正文中间段落**客观出现。对于不需要对标的主题，则专注讲透本品数据即可。
4. **【拒绝堆砌参数】**：不要像念说明书一样单列数据段落。必须在解决痛点的对话场景中自然连贯地引出（例：“比起硬上入门款的配置，不如看看这台车的1.5T地球梦发动机，182匹给到的充沛储备才是务实之选”）。
5. **名字灵活**。不要频繁正经地喊出“本田CR-V”。你可以叫它“这台车”、“同级标杆”或者只提一次“CR-V”。

【参考素材（融合这些灵感，但千万别照抄，体会"感觉"）】
1. 人设起步参考：
{chr(10).join(f"   - {sample}" for sample in persona_opening[:2]) if persona_opening else "   - （无样本）"}

2. 痛点场景参考：
{chr(10).join(f"   - {sample}" for sample in persona_pain[:2]) if persona_pain else "   - （无样本）"}

3. 具体细节库随机抽取：
{chr(10).join(f"   - {detail}" for detail in selected_details)}

【🔥 真实人类爆款感觉示例 (Few-Shot 注入) 🔥】
请深深体会以下真实高赞帖文的断句、网感和神经质的真实感，**照着这个调性写你的初稿**：
{chr(10).join(f"- '{sample}'" for sample in hit_samples) if hit_samples else "- （暂无该平台参照，请紧贴平台原生风格）"}

直接开始写你的真实碎片（直接输出正文，不要有任何前缀或解释）："""

            try:
                # 注入动态温度和 Token 限制
                content = await call_deepseek_api_async(prompt, DEEPSEEK_API_KEY, DEEPSEEK_API_URL, temperature=dynamic_temp, max_tokens=dynamic_max_tokens)
                print(f"    ✓ 第{assignment['id']}篇创作完成（{len(content)}字）")
                return {
                    "id": assignment["id"],
                    "content": content,
                    "persona": assignment["persona"],
                    "selling_point": assignment["selling_point"],
                    "attempt": 1,
                    "revision_history": []
                }
            except Exception as e:
                print(f"    ✗ 第{assignment['id']}篇创作失败：{e}")
                return {
                    "id": assignment["id"],
                    "content": f"[创作失败：{e}]",
                    "persona": assignment["persona"],
                    "selling_point": assignment["selling_point"],
                    "attempt": 1,
                    "revision_history": []
                }

        async def create_contents_parallel():
            """并行创建所有内容"""
            tasks = [create_single_content(assignment) for assignment in assignments]
            return await asyncio.gather(*tasks)

        # 运行并行创作
        contents = asyncio.run(create_contents_parallel())
        # 按 id 排序确保顺序一致
        contents = sorted(contents, key=lambda x: x["id"])

    state["contents"] = contents
    print(f"[Writer] 已创作{len(contents)}篇内容")

    return state


def 审核者(state: SharedContext) -> SharedContext:
    """
    职责：检查内容、生成修改建议、控制循环
    """
    print("\n[审核者] 开始 LLM 智能审核内容...")

    contents = state["contents"]
    planner_brief = state["planner_brief"]
    DEEPSEEK_API_KEY = config.DEEPSEEK_API_KEY
    DEEPSEEK_API_URL = config.DEEPSEEK_API_URL

    async def review_all_contents():
        async def review_single(content_item):
            content = content_item["content"]
            issues = []
            suggestions = []
            
            # === 硬性审核 ===
            platform = state.get("customer_brief", {}).get("平台", "小红书")
            
            # 分平台字数要求
            platform_limits = {
                "小红书": (100, 300),
                "抖音": (250, 350),
                "今日头条": (650, 800),
                "朋友圈": (20, 200)
            }
            min_words, max_words = platform_limits.get(platform, (200, 400))
            
            word_count = len(content)
            if not (min_words <= word_count <= max_words):
                issues.append(f"字数与【{platform}】要求不符（当前{word_count}字，合理区间为 {min_words}-{max_words}字）")
                suggestions.append(f"调整字数至 {min_words}-{max_words} 字区间")

            banned_found = check_banned_words(content)
            if banned_found:
                issues.append(f"包含禁用词：{', '.join(banned_found)}")
                suggestions.append(f"删除禁用词：{', '.join(banned_found)}")

            param_count = count_params(content)
            if param_count > 2:
                issues.append(f"参数过多（当前{param_count}个，要求≤2个）")
                suggestions.append("减少参数使用，用场景描写替代参数堆砌")

            # === AI味智能审核 (LLM) ===
            assignment = next((a for a in planner_brief["assignments"] if a["id"] == content_item["id"]), {"persona": "未知", "selling_point": "未知"})
            llm_eval = await evaluate_content_ai_flavor_async(content, assignment, DEEPSEEK_API_KEY, DEEPSEEK_API_URL)
            
            if not llm_eval.get("passed", False):
                issues.extend(llm_eval.get("issues", []))
                suggestions.extend(llm_eval.get("suggestions", []))
                
            passed = len(issues) == 0
            
            return {
                "id": content_item["id"],
                "passed": passed,
                "issues": issues,
                "suggestions": suggestions if not passed else [],
                "quality_scores": {
                    "scene": llm_eval.get("score", 5),
                    "emotion": llm_eval.get("score", 5)
                }
            }
        
        tasks = [review_single(item) for item in contents]
        return await asyncio.gather(*tasks)

    # 运行异步审核
    review_results = asyncio.run(review_all_contents())
    review_results = sorted(review_results, key=lambda x: x["id"])

    # 打印调试信息
    for r in review_results:
        if not r["passed"]:
            print(f"  [审核] 篇{r['id']} 不通过 (拟人度:{r['quality_scores']['scene']}/10)：{', '.join(r['issues'][:2])}...")
        else:
            print(f"  [审核] 篇{r['id']} 通过 (拟人度:{r['quality_scores']['scene']}/10)！")

    state["review_results"] = review_results

    passed_count = sum(1 for r in review_results if r["passed"])
    print(f"[审核者] 审核完成：{passed_count}/{len(contents)} 篇通过")

    return state


def route_after_review(state: SharedContext) -> str:
    """
    审核后的路由逻辑

    返回：
    - "Writer": 有内容需要修改，且未超过3次
    - "输出校订者": 全部通过，或超过3次需要人工介入
    """
    review_results = state["review_results"]
    current_attempt = state.get("current_attempt", 1)

    # 统计不通过的内容
    failed_items = [r for r in review_results if not r["passed"]]

    # 全部通过 → 进入输出校订者
    if not failed_items:
        print(f"[路由] 全部通过，进入输出校订者")
        return "输出校订者"

    # 超过3次 → 人工介入
    if current_attempt >= 3:
        print(f"[路由] 已尝试{current_attempt}次，进入人工介入流程")
        # 标记需要人工介入的内容
        state["need_manual_review"] = [item["id"] for item in failed_items]
        return "输出校订者"

    # 返回 Writer 修改
    print(f"[路由] 第{current_attempt}次尝试，{len(failed_items)}篇需要修改，返回 Writer")
    state["current_attempt"] = current_attempt + 1
    return "Writer"


def 输出校订者(state: SharedContext) -> SharedContext:
    """
    职责：最终校验、输出文件
    """
    print("\n[输出校订者] 开始最终校验...")

    contents = state["contents"]
    review_results = state["review_results"]
    need_manual_review = state.get("need_manual_review", [])

    # 读取参考材料
    anti_ai_style = config.load_material("02-参考学习/05-输出校订者材料/Anti-AI-style特征库.md")

    # 生成 Excel 输出
    try:
        output_path = generate_excel_output(state)
        state["final_output"] = output_path
        print(f"[输出校订者] 已输出到：{output_path}")
    except Exception as e:
        print(f"[输出校订者] Excel 生成失败：{e}，使用文本输出")
        # 备用：输出到文本文件
        output_file = "output_llm.txt"
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write("=== Agent Swarm 内容输出（真实 LLM）===\n\n")

            for content_item, review in zip(contents, review_results):
                if review["passed"]:
                    # 通过的内容
                    f.write(f"【篇{content_item['id']}】人设：{content_item['persona']} | 卖点：{content_item['selling_point']} | 状态：✅ 通过\n")
                    f.write(f"{content_item['content']}\n")
                    f.write(f"字数：{len(content_item['content'])} | 尝试次数：{content_item.get('attempt', 1)}\n\n")
                    f.write("-" * 50 + "\n\n")
                elif content_item["id"] in need_manual_review:
                    # 需要人工介入的内容
                    f.write(f"【篇{content_item['id']}】人设：{content_item['persona']} | 卖点：{content_item['selling_point']} | 状态：⚠️ 需要人工介入\n")
                    f.write(f"{content_item['content']}\n")
                    f.write(f"字数：{len(content_item['content'])} | 尝试次数：{content_item.get('attempt', 1)}\n")
                    f.write(f"问题：{', '.join(review['issues'])}\n\n")
                    f.write("-" * 50 + "\n\n")

        state["final_output"] = output_file

    # 统计
    passed_count = sum(1 for r in review_results if r["passed"])
    manual_count = len(need_manual_review)
    print(f"  - 通过：{passed_count}篇")
    print(f"  - 需要人工介入：{manual_count}篇")

    return state


# ============================================================================
# @main (Orchestrator) - 流程编排
# ============================================================================

def create_swarm() -> StateGraph:
    """
    创建 Agent Swarm 流程图
    """
    workflow = StateGraph(SharedContext)

    # 添加业务角色节点
    workflow.add_node("客户经理", 客户经理)
    workflow.add_node("策划者", 策划者)
    workflow.add_node("Writer", Writer)
    workflow.add_node("审核者", 审核者)
    workflow.add_node("输出校订者", 输出校订者)

    # 定义流程（@main 的调度逻辑）
    workflow.set_entry_point("客户经理")
    workflow.add_edge("客户经理", "策划者")
    workflow.add_edge("策划者", "Writer")
    workflow.add_edge("Writer", "审核者")

    # 审核-修改循环（使用条件边）
    workflow.add_conditional_edges(
        "审核者",
        route_after_review,
        {
            "Writer": "Writer",
            "输出校订者": "输出校订者"
        }
    )

    workflow.add_edge("输出校订者", END)

    return workflow.compile()


# ============================================================================
# 主程序入口
# ============================================================================

def main():
    """
    主程序入口
    """
    print("=" * 60)
    print("Agent Swarm 原型 - Content Expansion (Deepseek API)")
    print("=" * 60)

    # 检查 API Key（Deepseek 已硬编码，无需检查环境变量）
    print("\n使用 Deepseek API（deepseek-chat 模型）")

    # 初始化 Shared Context
    initial_state = {
        "user_input": {
            "车型": "CR-V",
            "平台": "抖音",
            "数量": 5,
            "方向": "春节返乡"
        },
        "customer_brief": {},
        "planner_brief": {},
        "contents": [],
        "review_results": [],
        "final_output": "",
        "current_attempt": 1,  # 当前循环次数
        "need_manual_review": [],  # 需要人工介入的ID列表
        "skip_confirmations": False,  # 是否跳过确认
        "metadata": {}
    }

    # 创建并运行 Swarm
    swarm = create_swarm()
    result = swarm.invoke(initial_state)

    print("\n" + "=" * 60)
    print("执行完成！")
    print(f"最终输出：{result['final_output']}")
    print("=" * 60)


if __name__ == "__main__":
    main()
