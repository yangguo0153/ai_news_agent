#!/usr/bin/env python3
"""
GEO效果监测Excel模板生成器
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

def find_repo_root(start: Path) -> Path:
    for path in [start, *start.parents]:
        if (path / ".git").exists():
            return path
    return start

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "GEO效果监测"

# 定义样式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 定义字段（精简版10字段）
fields = [
    ("检测日期", 12, "YYYY-MM-DD格式"),
    ("AI平台", 14, "豆包/文心一言/DeepSeek/通义千问"),
    ("查询关键词", 35, "实际输入AI的完整问题"),
    ("关键词类型", 12, "品牌词/场景词/对比词/痛点词"),
    ("是否露出", 10, "是/否"),
    ("露出位置", 10, "1-10，未露出填0"),
    ("露出类型", 14, "首选推荐/推荐型/列举型/仅提及"),
    ("我方胜出", 10, "是/否/无竞品"),
    ("露出原文", 40, "AI提及LS9的原话（50字内）"),
    ("证据截图", 25, "截图链接或文件名"),
]

# 写入表头
for col, (field_name, width, _) in enumerate(fields, 1):
    cell = ws.cell(row=1, column=col, value=field_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width

# 设置行高
ws.row_dimensions[1].height = 30

# 添加数据验证（下拉选项）
# AI平台
dv_platform = DataValidation(type="list", formula1='"豆包,文心一言,DeepSeek,通义千问"', allow_blank=True)
dv_platform.error = "请从下拉列表中选择"
dv_platform.errorTitle = "无效输入"
ws.add_data_validation(dv_platform)
dv_platform.add("B2:B1000")

# 关键词类型
dv_keyword_type = DataValidation(type="list", formula1='"品牌词,场景词,对比词,痛点词"', allow_blank=True)
ws.add_data_validation(dv_keyword_type)
dv_keyword_type.add("D2:D1000")

# 是否露出
dv_exposed = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
ws.add_data_validation(dv_exposed)
dv_exposed.add("E2:E1000")

# 露出类型（删除"未露出"选项，因为已有"是否露出"字段）
dv_expose_type = DataValidation(type="list", formula1='"首选推荐,推荐型,列举型,仅提及"', allow_blank=True)
ws.add_data_validation(dv_expose_type)
dv_expose_type.add("G2:G1000")

# 我方胜出（位置从K改为H）
dv_win = DataValidation(type="list", formula1='"是,否,无竞品"', allow_blank=True)
ws.add_data_validation(dv_win)
dv_win.add("H2:H1000")

# 添加示例数据（精简版10字段）
sample_data = [
    ["2026-01-28", "豆包", "50万纯电轿车推荐", "场景词", "是", 2, "推荐型", "否", "智己LS9续航超700km，智驾能力领先同级", "screenshot_01.png"],
    ["2026-01-28", "DeepSeek", "智己LS9怎么样", "品牌词", "是", 1, "首选推荐", "无竞品", "LS9是智己旗舰轿车，综合实力出众", "screenshot_02.png"],
    ["2026-01-28", "文心一言", "家用大空间电车", "场景词", "否", 0, "", "", "", ""],
]

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# 创建字段说明Sheet
ws2 = wb.create_sheet("字段说明")
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 50
ws2.column_dimensions['C'].width = 40

# 字段说明表头
for col, header in enumerate(["字段名", "填写说明", "甲方价值"], 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 字段说明内容（精简版10字段）
field_docs = [
    ("检测日期", "YYYY-MM-DD格式", "时间趋势追踪"),
    ("AI平台", "豆包/文心一言/DeepSeek/通义千问", "分平台看效果分布"),
    ("查询关键词", "实际输入AI的完整问题", "追溯测了哪些词"),
    ("关键词类型", "品牌词/场景词/对比词/痛点词", "分类统计覆盖率"),
    ("是否露出", "是/否", "【核心KPI】露出率计算"),
    ("露出位置", "1-10的数字，未露出填0", "【核心KPI】平均排名计算"),
    ("露出类型", "首选推荐 > 推荐型 > 列举型 > 仅提及", "露出质量分级"),
    ("我方胜出", "是=排名优于竞品，否=排名落后，无竞品=回答中只有我方", "【核心KPI】竞品胜率"),
    ("露出原文", "AI提及智己LS9的原话，50字内", "证据留存，可追溯"),
    ("证据截图", "截图链接或文件名", "防扯皮，可追溯"),
]

for row_idx, (field, desc, value) in enumerate(field_docs, 2):
    ws2.cell(row=row_idx, column=1, value=field)
    ws2.cell(row=row_idx, column=2, value=desc)
    ws2.cell(row=row_idx, column=3, value=value)

# 创建关键词类型说明Sheet
ws3 = wb.create_sheet("关键词类型定义")
ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 45

for col, header in enumerate(["类型", "定义", "示例"], 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill

keyword_types = [
    ("品牌词", "直接包含品牌/产品名", "智己LS9怎么样、智己LS9续航"),
    ("场景词", "购车场景/需求描述", "50万纯电轿车推荐、家用大空间电车"),
    ("对比词", "明确对比多个品牌", "智己LS9和蔚来ET7选哪个"),
    ("痛点词", "用户痛点/需求导向", "续航长不焦虑的电车、智能驾驶好的车"),
]

for row_idx, (ktype, definition, example) in enumerate(keyword_types, 2):
    ws3.cell(row=row_idx, column=1, value=ktype)
    ws3.cell(row=row_idx, column=2, value=definition)
    ws3.cell(row=row_idx, column=3, value=example)

# 保存文件
repo_root = find_repo_root(Path(__file__).resolve())
output_path = repo_root / "公司业务/业务文档/GEO效果监测模板.xlsx"
output_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(output_path)
print(f"✅ Excel模板已生成: {output_path}")
